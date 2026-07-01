from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from scheduler import get_month_days, normalize_schedule_df


HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
FRIDAY_FILL = PatternFill("solid", fgColor="D9EAF7")
SATURDAY_FILL = PatternFill("solid", fgColor="E2F0D9")
SUNDAY_FILL = PatternFill("solid", fgColor="FCE4D6")
RED_FONT = Font(color="C00000")
BLACK_FONT = Font(color="000000")
TITLE_FONT = Font(size=14, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
BLOCK_DAYS = 14


def weekday_fill(weekday_index: int) -> PatternFill | None:
    if weekday_index == 4:
        return FRIDAY_FILL
    if weekday_index == 5:
        return SATURDAY_FILL
    if weekday_index == 6:
        return SUNDAY_FILL
    return None


def write_schedule_block(ws, schedule_df: pd.DataFrame, days: list[dict], start_row: int) -> int:
    ws.cell(row=start_row, column=1, value="序号")
    ws.cell(row=start_row, column=2, value="姓名")
    ws.cell(row=start_row + 1, column=1, value="")
    ws.cell(row=start_row + 1, column=2, value="")

    for column_index, item in enumerate(days, start=3):
        ws.cell(row=start_row, column=column_index, value=item["label"])
        ws.cell(row=start_row + 1, column=column_index, value=item["weekday"])

    for row_offset, record in enumerate(schedule_df.to_dict("records"), start=2):
        row_index = start_row + row_offset
        ws.cell(row=row_index, column=1, value=record["序号"])
        ws.cell(row=row_index, column=2, value=record["姓名"])
        for column_index, item in enumerate(days, start=3):
            value = record.get(item["date"], "")
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.font = RED_FONT if value in ["休息", "班/休"] else BLACK_FONT

    block_last_row = start_row + 1 + len(schedule_df)
    block_last_column = 2 + len(days)
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=block_last_row,
        min_col=1,
        max_col=block_last_column,
    ):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    for column_index in range(1, block_last_column + 1):
        ws.cell(row=start_row, column=column_index).fill = HEADER_FILL
        ws.cell(row=start_row + 1, column=column_index).fill = HEADER_FILL

    for column_index, item in enumerate(days, start=3):
        fill = weekday_fill(int(item["weekday_index"]))
        if fill:
            for row_index in range(start_row, block_last_row + 1):
                ws.cell(row=row_index, column=column_index).fill = fill

    return block_last_row


def export_schedule_excel(schedule_df: pd.DataFrame, year: int, month: int) -> bytes:
    days = get_month_days(year, month)
    schedule_df = normalize_schedule_df(schedule_df, year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = "排班表"

    max_column = 2 + min(BLOCK_DAYS, len(days))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    title_cell = ws.cell(row=1, column=1, value=f"{year}年{month}月科室排班表")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    start_row = 3
    for start in range(0, len(days), BLOCK_DAYS):
        block_days = days[start:start + BLOCK_DAYS]
        last_row = write_schedule_block(ws, schedule_df, block_days, start_row)
        start_row = last_row + 2

    for column_index in range(1, max_column + 1):
        ws.column_dimensions[get_column_letter(column_index)].width = 6
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "C5"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
